from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
EXCEL_FILE = "data.xlsx"

# Ensure Excel file exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Amount", "Room No"])
    wb.save(EXCEL_FILE)

@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        amount = request.form["amount"]
        room_no = request.form["room_no"]

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([amount, room_no])
        wb.save(EXCEL_FILE)

        return redirect(url_for("form"))
    return render_template("form.html")

@app.route("/view")
def view():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    data = list(ws.values)  # List of rows
    return render_template("view.html", data=data)

if __name__ == "__main__":
    app.run(debug=True)
